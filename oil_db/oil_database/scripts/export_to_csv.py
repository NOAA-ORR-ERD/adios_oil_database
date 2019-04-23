import sys
import os
import logging

import numpy as np

from openpyxl import Workbook

import unit_conversion as uc

from oil_database.util.db_connection import connect_modm
from oil_database.util.settings import file_settings, default_settings

logger = logging.getLogger(__name__)


def export_to_csv_cmd(argv=sys.argv):
    logger.setLevel(logging.INFO)
    logger.addHandler(logging.StreamHandler())

    if len(argv) < 2:
        export_to_csv_usage(argv)

    export_file = argv[1]

    if len(argv) >= 3:
        settings = file_settings(argv[2])
    else:
        settings = default_settings()

    try:
        export_to_csv(settings, export_file)
    except Exception:
        print "{0}() FAILED\n".format(export_to_csv.__name__)
        raise


def export_to_csv_usage(argv):
    cmd = os.path.basename(argv[0])

    print('usage: {0} <export_file> [config_file]\n'
          '(example: "{0} ec_export.csv")'.format(cmd))

    sys.exit(1)


def export_to_csv(settings, export_file):
    logger.info('connect_modm()...')
    connect_modm(settings)

    wb = Workbook()

    wb.remove(wb.active)
    dvis_sheet, api_sheet, hopane_sheet = add_wb_sheets(wb)

    logger.info('add_all_ec_records()...')
    add_all_ec_records(dvis_sheet, api_sheet, hopane_sheet)

    wb.save(export_file)


def add_wb_sheets(wb):
    dvis_sheet = wb.create_sheet("Viscosities", 0)
    api_sheet = wb.create_sheet("APIs", 1)
    hopane_sheet = wb.create_sheet("Hopanes", 2)

    return dvis_sheet, api_sheet, hopane_sheet


def add_all_ec_records(dvis_sheet, api_sheet, hopane_sheet):
    '''
        Note: We can't import any PyMODM models until there is a connection,
              so we import it here.
              We will assume the connect() has already happened prior to this
              function.
    '''
    from oil_database.models.ec_imported_rec import ECImportedRecord

    dvis_counts = []
    api_counts = []
    hopane_counts = []

    for idx, ec_rec in enumerate(ECImportedRecord.objects.all()):
        dvis_counts.append(add_dvis_cells(dvis_sheet, idx, ec_rec))
        api_counts.append(add_api_cells(api_sheet, idx, ec_rec))
        hopane_counts.append(add_hopane_cells(hopane_sheet, idx, ec_rec))

    add_dvis_field_header(dvis_sheet, max(dvis_counts))
    add_api_field_header(api_sheet, max(api_counts))
    add_hopane_field_header(hopane_sheet, max(hopane_counts))

    add_sorting_to_sheet(dvis_sheet, idx, max(dvis_counts) * 2 + 2)
    add_sorting_to_sheet(api_sheet, idx, max(api_counts) * 2 + 1)
    add_sorting_to_sheet(hopane_sheet, idx, max(hopane_counts) * 2 + 1)


def add_dvis_cells(sheet, row, ec_rec):
    '''
        Note: This is an attempt at representing the format of the spreadsheet
              that RobertJ has supplied.
              The first sheet contains the following columns:
              - Oil Name
              - Temperature [C]
              - number of pts
              - % evaporated
              - viscosity (mPa.s)
              - % evaporated
              - viscosity (mPa.s)
              - % evaporated
              - viscosity (mPa.s)
              - % evaporated
              - viscosity (mPa.s)

        Note: It looks like we just want the viscosities measured at 0C
        Note: the number of points seems to indicate that there can be a
              variable number of columns depending upon how many viscosities
              are represented in the source record.
    '''
    ret = []
    ret.append(ec_rec.oil_name)

    dvis = ([d for d in ec_rec.dvis
             if np.isclose(d.ref_temp_k, 273.15, rtol=0.001)])

    ret.append(0.0)
    ret.append(len(dvis))

    for d in dvis:
        ret.extend((d.weathering,
                    uc.convert('Dynamic Viscosity', 'kg/(m s)', 'mPa s',
                               d.kg_ms)))

    for i, d in enumerate(ret):
        # we start on the second row
        sheet.cell(row=row + 2, column=i + 1).value = d

    return len(dvis)


def add_api_cells(sheet, row, ec_rec):
    '''
        Note: This is an attempt at representing the format of the spreadsheet
              that RobertJ has supplied.
              The second sheet contains the following columns:
              - Oil Name
              - number of pts
              - % evaporated
              - API
              - % evaporated
              - API
              - % evaporated
              - API
              - % evaporated
              - API

        Note: the number of points seems to indicate that there can be a
              variable number of columns depending upon how many viscosities
              are represented in the source record.
    '''
    ret = []

    ret.append(ec_rec.oil_name)
    ret.append(len(ec_rec.api))

    # We assume that the number of densities will be the same as
    # the number of APIs.  Good assumption?
    # Should API be a class object with its own weathering property?
    for a, d in zip(ec_rec.api, ec_rec.densities):
        ret.extend((d.weathering, a))

    for i, d in enumerate(ret):
        # we start on the second row
        sheet.cell(row=row + 2, column=i + 1).value = d

    return len(ec_rec.api)


def add_hopane_cells(sheet, row, ec_rec):
    '''
        Note: This is an attempt at representing the format of the spreadsheet
              that RobertJ has supplied.
              The third sheet contains the following columns:
              - Oil Name
              - number of pts
              - % evaporated
              - Hopane (H30) (ug/g)
              - % evaporated
              - Hopane (H30) (ug/g)
              - % evaporated
              - Hopane (H30) (ug/g)
              - % evaporated
              - Hopane (H30) (ug/g)
    '''
    ret = []

    ret.append(ec_rec.oil_name)
    ret.append(len(ec_rec.biomarkers))

    for b in ec_rec.biomarkers:
        ret.extend((b.weathering,
                    b.hopane_ug_g))

    for i, d in enumerate(ret):
        # we start on the second row
        sheet.cell(row=row + 2, column=i + 1).value = d

    return len(ec_rec.biomarkers)


def add_dvis_field_header(sheet, num_groups):
    ret = []

    ret.extend(('Oil Name',
                'Temperature [C]',
                'number of pts'))

    for _ in range(num_groups):
        ret.extend(('% evaporated',
                    'viscosity (mPa.s)'))

    for i, d in enumerate(ret):
        sheet.cell(row=1, column=i + 1).value = d


def add_api_field_header(sheet, num_groups):
    ret = []

    ret.extend(('Oil Name',
                'number of pts'))

    for _ in range(num_groups):
        ret.extend(('% evaporated',
                    'API'))

    for i, d in enumerate(ret):
        sheet.cell(row=1, column=i + 1).value = d


def add_hopane_field_header(sheet, num_groups):
    ret = []

    ret.extend(('Oil Name',
                'number of pts'))

    for _ in range(num_groups):
        ret.extend(('% evaporated',
                    'Hopane (H30) (ug/g)'))

    for i, d in enumerate(ret):
        sheet.cell(row=1, column=i + 1).value = d


def add_sorting_to_sheet(sheet, row_count, column_count):
    sheet.auto_filter.ref = ('A1:{}{}'
                             .format(excel_column_name(column_count),
                                     row_count))

    sheet.auto_filter.add_sort_condition('A2:A{}'.format(row_count))


def excel_column_name(idx):
    ret = []
    if idx >= 26:
        ret.append(chr(0x41 + (idx / 26 - 1)))

    ret.append(chr(0x41 + (idx % 26)))

    return ''.join(ret)

















