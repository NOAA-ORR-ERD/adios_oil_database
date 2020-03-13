#!/usr/bin/env python

import sys
import os
import logging
from openpyxl import load_workbook

from ..import field_name_sluggify

from pprint import PrettyPrinter
pp = PrettyPrinter(indent=2, width=120)

custom_slugify = field_name_sluggify

logger = logging.getLogger(__name__)


class ExxonDataReader:
    """
    reads the Exxon Excel files, and returns records as simple nested
    data structures:

    Essentially the file as a raw table
    """
    def __init__(self, data_index_file, data_dir=None):
        """
        Initialize a reader for the Exxon Data

        :param data_dir: directory with one or more Exxon Assay Excel files

        :param data_index_file: name of index file -- mapping oil names
                                to files
        """
        self.data_dir = data_dir
        if self.data_dir is None:
            self.data_dir = os.path.dirname(data_index_file)

        self.index = self._read_index(data_index_file)

    def _read_index(self, data_index_file):
        with open(data_index_file) as indexfile:
            header = indexfile.readline().strip()
            if header.split() != ['oil_name', 'file']:
                raise ValueError(f"This: {data_index_file}\n"
                                 "does not look like an Exxon Assay Index file")
            index = []

            for line in indexfile:
                line = line.strip()
                if not line:
                    continue

                try:
                    oil_name, filename = line.split("\t")
                except ValueError:
                    raise ValueError(f"There is something wrong with this line"
                                     "in the index file:\n"
                                     "{line}")
                index.append((oil_name,
                              os.path.join(self.data_dir, filename)))

            return index

    def get_records(self):
        # an empty list for now -- just so we can test it :-)
        for oilname, filename in self.index:
            yield (oilname, self.read_excel_file(filename))

    @staticmethod  # make it easier to test on its own
    def read_excel_file(filename):
        """
        The code that reads an excel file, and returns what's in it
        as a nested list.

        Note: you'd think this would be a single call in openpyxl, but I
              couldn't find it
        """

        wb = load_workbook(filename, data_only=True)

        sheetnames = wb.sheetnames

        if len(sheetnames) != 1:
            raise ValueError(f'file: {filename} does not have '
                             'one and only one sheet')

        sheet = wb[sheetnames[0]]

        return [[c.value for c in row] for row in sheet.rows]
