#!/usr/bin/env python
import os
from pathlib import Path
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExxonDataReader:
    """
    reads the Exxon Excel files, and returns records as simple nested
    data structures:

    Essentially the file as a raw table
    """
    def __init__(self, data_index_file, data_dir=None):
        """
        Initialize a reader for the Exxon Data

        :param data_dir: directory with one or more Exxon Assay Excel files

        :param data_index_file: name of index file -- mapping oil names
                                to files
        """
        self.data_dir = data_dir
        if self.data_dir is None:
            self.data_dir = Path(os.path.dirname(data_index_file))

        self.index = self._read_index(data_index_file)

    def _read_index(self, data_index_file):
        with open(data_index_file, encoding="utf-8") as indexfile:
            header = indexfile.readline().strip()
            if header.split() != ['oil_name', 'file']:
                raise ValueError(f"This: {data_index_file}\n"
                                 "does not look like an Exxon Assay "
                                 "Index file")
            index = []

            for line in indexfile:
                line = line.strip()
                if not line:
                    continue

                try:
                    oil_name, filename = line.split("\t")
                except ValueError:
                    raise ValueError(f"There is something wrong with this line"
                                     "in the index file:\n"
                                     "{line}")
                index.append({
                    'name': oil_name,
                    'path': self.data_dir / filename
                })

            return index

    def get_records(self):
        # an empty list for now -- just so we can test it :-)
        for i in self.index:
            yield (i['name'], self.read_excel_file(i['path']))

    @staticmethod  # make it easier to test on its own
    def read_excel_file(filename):
        """
        The code that reads an excel file, and returns what's in it
        as a nested list.

        Note: you'd think this would be a single call in openpyxl, but I
              couldn't find it
        Note: Recently, a new format for the Exxon assays has appeared, which
              contains 2 sheets.  For the moment, we can fail to parse the
              new formatted information, but reading the first sheet should not
              fail.
        """
        print("reading:", filename)
        wb = load_workbook(filename, data_only=True)

        sheetnames = wb.sheetnames

        if len(sheetnames) < 1:
            raise ValueError(f'file: {filename} does not contain '
                             'any sheets')

        sheets = [[[c.value for c in r]
                   for r in s.rows]
                  for s in [wb[n] for n in sheetnames]
                  ]

        return sheets
